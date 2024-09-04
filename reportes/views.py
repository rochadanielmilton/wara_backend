from django.db import connection, models
from django.forms import FloatField
import openpyxl
from django.db.models import Avg, F, Q,Prefetch
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from backend_mmaya.views import SoftDeleteModelViewSet
from .serializers import *
# from administracion.models import Realizacion
from django.http import JsonResponse
from rest_framework.response import Response
from django.db.models.functions import Cast
from django.db.models import Count, Sum,F,Sum, Avg, F
from django.views.generic import View
from programas.models import*
from babel.numbers import format_decimal
from rest_framework.views import APIView
from parametros.models import Departamento, Municipio, Provincia
from rest_framework import generics, permissions
from rest_framework import viewsets,status
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
import os
from django.core.files.images import ImageFile
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.utils.timezone import now
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

class ListaProyectosView(SoftDeleteModelViewSet):

    queryset = Proyectos.objects.all()  
    serializer_class = ProyectoSerializer
class ProyectoPorGestionView(View):
    def get(self, request, *args, **kwargs):
        sql_query = """
            SELECT 
                TO_CHAR(TO_DATE(SUBSTRING(p.fecha_conclusion, '\d{1,2}/\d{1,2}/\d{4}'), 'MM/DD/YYYY'), 'YYYY') AS gestion,
                COUNT(*) AS cantidad_proyectos,
                SUM(r.total_inversion) AS total_inversion
            FROM 
                proyectos p
            INNER JOIN 
                realizaciones r 
            ON 
                r.proyecto_id = p.id
            WHERE   
                p.fecha_conclusion !~ '^[A-Za-z]' 
                AND (SELECT (REGEXP_MATCHES(p.fecha_conclusion, '\d{1,2}'))[1]::integer) <= 12
            GROUP BY 
                TO_CHAR(TO_DATE(SUBSTRING(p.fecha_conclusion, '\d{1,2}/\d{1,2}/\d{4}'), 'MM/DD/YYYY'), 'YYYY')
            ORDER BY 
                TO_CHAR(TO_DATE(SUBSTRING(p.fecha_conclusion, '\d{1,2}/\d{1,2}/\d{4}'), 'MM/DD/YYYY'), 'YYYY');
        """

        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            rows = cursor.fetchall()

            result_list = []
            for row in rows:
                result_dict = {
                    'gestion': row[0],
                    'cantidad_proyectos': row[1],
                    'total_inversion': row[2]
                }
                result_list.append(result_dict)

            return JsonResponse(result_list, safe=False)

class ProgramasProyectosInversionPorDeptoProvMunView(APIView):
    def get(self, request, *args, **kwargs):
        departamento = request.GET.get('departamento_id')
        provincia = request.GET.get('provincia_id')
        estado_id = request.GET.get('estado_id')
    
        sql_por_departamento = """
        SELECT d.id, d.nombre, subQuery.cantidad_proyectos, subQuery.total_inversion, subQuery.cantidad_programas
        FROM departamentos d
        INNER JOIN (
            SELECT rd.departamento_id,
                   COUNT(*) AS cantidad_proyectos,
                   SUM(r.total_inversion) AS total_inversion,
                   COUNT(DISTINCT p.programa_id) AS cantidad_programas
            FROM realizaciones r
            INNER JOIN proyectos p ON r.proyecto_id = p.id
            INNER join realizacion_departamentos rd on r.id =rd.realizacion_id
            WHERE (%s IS NULL OR r.estado_id=%s)
            GROUP BY rd.departamento_id
        ) subQuery ON d.id=subQuery.departamento_id
        ORDER BY d.nombre;
        """
        
        sql_por_provincia = """
        SELECT p.id, p.nombre, subQuery.cantidad_proyectos, subQuery.total_inversion, subQuery.cantidad_programas
        FROM provincias p
        INNER JOIN (
            SELECT rp.provincia_id,
                   COUNT(*) AS cantidad_proyectos,
                   SUM(r.total_inversion) AS total_inversion,
                   COUNT(DISTINCT p.programa_id) AS cantidad_programas
            FROM realizaciones r
            INNER JOIN proyectos p ON r.proyecto_id = p.id
            INNER join realizacion_departamentos rd on r.id =rd.realizacion_id
            INNER join realizacion_provincias rp on r.id =rp.realizacion_id
            WHERE (%s IS NULL OR r.estado_id=%s) AND rd.departamento_id=%s
            GROUP BY rp.provincia_id
        ) subQuery ON p.id=subQuery.provincia_id
        ORDER BY p.nombre;
        """

        sql_por_municipio = """
        SELECT m.id, m.nombre, subQuery.cantidad_proyectos, subQuery.total_inversion, subQuery.cantidad_programas
        FROM municipios m
        INNER JOIN (
            SELECT rm.municipio_id,
                   COUNT(*) AS cantidad_proyectos,
                   SUM(r.total_inversion) AS total_inversion,
                   COUNT(DISTINCT p.programa_id) AS cantidad_programas
            FROM realizaciones r
            INNER JOIN proyectos p ON r.proyecto_id = p.id
            INNER JOIN realizacion_provincias rp ON r.id =rp.realizacion_id 
            INNER JOIN realizacion_municipios rm ON r.id =rm.realizacion_id 
            WHERE (%s IS NULL OR r.estado_id=%s) AND rp.provincia_id=%s
            GROUP BY rm.municipio_id
        ) subQuery ON m.id=subQuery.municipio_id
        ORDER BY m.nombre;
        """
        
        with connection.cursor() as cursor:
            if provincia is not None:
                cursor.execute(sql_por_municipio, (estado_id, estado_id, provincia))    
            elif departamento is not None:
                cursor.execute(sql_por_provincia, (estado_id, estado_id, departamento))    
            else:
                cursor.execute(sql_por_departamento, (estado_id, estado_id))
                
            rows = cursor.fetchall()

            result_list = []
            for row in rows:
                result_dict = {
                    'id': row[0],
                    'nombre': row[1],
                    'cantidad_proyectos': row[2],
                    'total_inversion': row[3],
                    'cantidad_programas': row[4]
                }
                result_list.append(result_dict)

            return JsonResponse(result_list, safe=False)

        
    
class datosPorSectoresView(APIView):
    def get(self, request, *args, **kwargs):
        departamento_id = request.query_params.get('departamento_id')
        provincia_id = request.query_params.get('provincia_id')
        municipio_id = request.query_params.get('municipio_id')
        estado_id = request.query_params.get('estado_id')

        consulta = {
            "departamento_id": departamento_id,
            "provincia_id": provincia_id,
            "municipio_id": municipio_id,
            "estado_id": estado_id
        }

        filter_criteria = {}

        if estado_id:
            filter_criteria['estado_id'] = estado_id

        if municipio_id:
            filter_criteria['realizacionmunicipios__municipio_id'] = municipio_id
        elif provincia_id:
            filter_criteria['realizacionprovincias__provincia_id'] = provincia_id
        elif departamento_id:
            filter_criteria['realizaciondepartamentos__departamento_id'] = departamento_id

        realizaciones_queryset = Realizaciones.objects.filter(**filter_criteria)

        agregacion_por_sector = realizaciones_queryset.values(
            sector_nombre=F('proyecto__sector_clasificador__nombre'),
            sector_id=F('proyecto__sector_clasificador__id')
        ).annotate(
            total_inversion=Sum('total_inversion'),
            cantidad_proyectos=Count('proyecto_id')
        ).order_by('sector_nombre')

        total_inversion_todos_los_sectores = realizaciones_queryset.aggregate(
            total_inversion=Sum('total_inversion')
        )['total_inversion']

        total_proyectos = realizaciones_queryset.count()

        response_data = {
            'total_inversion': total_inversion_todos_los_sectores,
            'total_numero_proyectos': total_proyectos,
            'inversion_por_sector': list(agregacion_por_sector),
            'consulta': consulta
        }

        return Response(response_data)



class InversionPorGobiernoView(APIView):
    def get(self, request, *args, **kwargs):
        gobierno = request.query_params.get('gobierno')
        if gobierno:
            realizaciones = Realizaciones.objects.filter(proyecto__gobierno=gobierno)
        else:
            realizaciones = Realizaciones.objects.all() 


        total_inversion = realizaciones.aggregate(total_inversion=Sum('total_inversion'))['total_inversion']
        total_inversion_formateado= format_decimal(total_inversion,locale='es_ES',format='#,##0.00')
        agregacion_por_departamento = realizaciones.values(departamento_nombre=F('departamento__nombre')).annotate(
            total_inversion=Sum('total_inversion'),
            cantidad_proyectos=Count('proyecto_id')
        )
        for item in agregacion_por_departamento:
            item['total_inversion'] =format_decimal(item['total_inversion'],locale='es_ES',format='#,##0.00')
        response_data = {
            'total_inversion': total_inversion_formateado,
            'inversion_por_departamentos': list(agregacion_por_departamento)
        }

        return Response(response_data)

class TotalInversionPorcentajes(APIView):
    permission_classes = [permissions.IsAuthenticated] 
    def get(self, request, *args, **kwargs):
        sql_query = """
            SELECT 
                subQuery.total_inversion,
                subQuery.total_inversion * (subQuery.avance_financiamiento / 100) AS avance_financiamiento,
                subQuery.total_inversion * (subQuery.avance_financiamiento_faltante / 100) AS avance_faltante
            FROM 
            (
                SELECT 
                    SUM(total_inversion) AS total_inversion,
                    ROUND(AVG(avance_financiamiento::integer), 2) AS avance_financiamiento,
                    (100 - ROUND(AVG(avance_financiamiento::integer), 2)) AS avance_financiamiento_faltante
                FROM 
                    realizaciones r
                WHERE 
                    avance_financiamiento !~ '^[A-Za-z]' 
                    AND avance_financiamiento::integer > 0 
                    AND total_inversion != total_inversion * avance_financiamiento::integer / 100
            ) subQuery;
            """

        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            rows = cursor.fetchall()
            result_list = []
            for row in rows:
                result_dict = {
                    'total_inversion': format_decimal(row[0],locale='es_ES',format='#,##0.00')if row[0] is not None else None,
                    'avance_financiamiento': format_decimal(row[1],locale='es_ES',format='#,##0.00')if row[1] is not None else None,
                    'avance_faltante': format_decimal(row[2],locale='es_ES',format='#,##0.00') if row[2] is not None else None,
                }
                result_list.append(result_dict)

        return JsonResponse(result_list, safe=False)

from django.http import JsonResponse
from django.db import connection
from rest_framework.views import APIView

from django.http import JsonResponse
from django.db import connection
from rest_framework.views import APIView

class ListProyectosPorProvinciaView(APIView):
    def get(self, request, *args, **kwargs):
        departamento_id = request.GET.get('departamento_id')
        provincia_id = request.GET.get('provincia_id')
        municipio_id = request.GET.get('municipio_id')
        estado = request.GET.get('estado_id')
        realizacion_id = request.GET.get('realizacion_id')
        search = request.GET.get('search', None)

        sql_query = """
        SELECT r.id, p.codigo_sisin, d.departamento_id, d2.nombre AS departamento, 
            pr.provincia_id, p2.nombre AS provincia, m.municipio_id, m2.nombre AS municipio,
            TRIM(p.nombre) AS nombre, p.programa_id, pro.sigla_prog_convenio AS programa,
            organizacion_id, org.sigla AS organizacion,
            p.ejecutor_id, e.nombre AS ejecutor, 
            tipo_id, tp.nombre AS tipo,
            p.estado_id, ep.nombre AS estado, TRIM(p.estado_detallado) AS estado_detallado,
            p.fecha_inicio, p.fecha_conclusion, r.total_inversion, r.avance_fisico, r.avance_financiamiento,
            p.gobierno  
        FROM proyectos p 
		INNER JOIN realizaciones r ON p.id = r.proyecto_id
		INNER JOIN realizacion_departamentos d ON r.id = d.realizacion_id
		INNER JOIN departamentos d2 ON d2.id = d.departamento_id 
		INNER JOIN realizacion_provincias pr ON r.id = pr.realizacion_id
		INNER JOIN provincias p2 ON p2.id = pr.provincia_id 
		INNER JOIN realizacion_municipios m ON r.id = m.realizacion_id
		INNER JOIN municipios m2 ON m2.id = m.municipio_id 
		INNER JOIN programas pro ON p.programa_id = pro.id
		INNER JOIN organizaciones_financieras org ON p.organizacion_id = org.id 
		INNER JOIN ejecutores e ON e.id = p.ejecutor_id  
		INNER JOIN tipos_proyecto tp ON tp.id = p.tipo_id  
		INNER JOIN estados2 ep ON ep.id = p.estado_id 
        WHERE (%s IS NULL OR ep.id = %s)
        """

        
        
        params = [estado, estado]

        if departamento_id:
            sql_query += " AND d.departamento_id = %s"
            params.append(departamento_id)
        if provincia_id:
            sql_query += " AND pr.provincia_id = %s"
            params.append(provincia_id)
        if municipio_id:
            sql_query += " AND m.municipio_id = %s"
            params.append(municipio_id)
        if realizacion_id:
            sql_query += " AND r.id = %s"
            params.append(realizacion_id)
        if search:
            sql_query+= "and (p.nombre ilike %s or pro.programas_proyectos ilike %s or pro.sigla_prog_convenio ilike %s)"
            params.extend(["%" + search + "%", "%" + search + "%", "%" + search + "%"])
 

        sql_query += " ORDER BY p.nombre;"

        with connection.cursor() as cursor:
            cursor.execute(sql_query, params)
            rows = cursor.fetchall()

            result_list = []
            for row in rows:
                result_dict = {
                    "id": row[0],
                    "codigo_sisin": row[1],
                    "departamento_id": row[2],
                    "departamento": row[3],
                    "provincia_id": row[4],
                    "provincia": row[5],
                    "municipio_id": row[6],
                    "municipio": row[7],
                    "nombre": row[8],
                    "programa_id": row[9],
                    "sigla": row[10],
                    "organizacion_id": row[11],
                    "organizacion": row[12],
                    "ejecutor_id": row[13],
                    "ejecutor": row[14],
                    "tipo_id": row[15],
                    "tipo": row[16],
                    "estado_id": row[17],
                    "estado": row[18],
                    "estado_detallado": row[19],
                    "fecha_inicio": row[20],
                    "fecha_conclusion": row[21],
                    "total_inversion": row[22],
                    "avance_fisico": row[23],
                    "avance_financiamiento": row[24],
                    "gobierno": row[25],
                }
                result_list.append(result_dict)

            return JsonResponse(result_list, safe=False)
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def proyectos_por_busqueda_pdf(request):
    template_path = 'proyectos-por-busqueda.html'
    base_url = f"{request.scheme}://{request.get_host()}"

    search = request.GET.get('search', '')
    queryset = Proyectos.objects.filter(is_deleted=False, estado_id=3)

    if search:
        queryset = queryset.filter(
            Q(nombre__icontains=search) | 
            Q(programa__sigla_prog_convenio__icontains=search)
        )
    else:
        return HttpResponse('No se encontro registros')

    # Use select_related to perform an SQL JOIN and include related Realizaciones data
    queryset = queryset.select_related('realizaciones')

    context = {
        'usuario':request.user,
        'base_url': base_url,
        'search':search,
        'current_datetime': now(),
        'proyectos': queryset,
    }

    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="proyectos.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse(f'Failed to generate PDF: {pisa_status.err}', status=500)

    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def proyecto_por_municipio_pdf(request):
    template_path = 'proyectos-por-municipio.html'
    base_url = f"{request.scheme}://{request.get_host()}"
    context = {
        'usuario':request.user,
        'base_url': base_url,
        'current_datetime': now(),
    }

    # Obtener los parámetros de la solicitud GET
    departamento_id = request.GET.get('departamento_id')
    provincia_id = request.GET.get('provincia_id')
    municipio_id = request.GET.get('municipio_id')
    estado_id = request.GET.get('estado_id')
    realizacion_id = request.GET.get('realizacion_id')
    proyectos = []
    avances_fisicos = []
    avances_financieros = []

    if estado_id == "0":
        estado_id = None

    # Asegurar que los parámetros departamento_id, provincia_id y municipio_id sean proporcionados
    if not (departamento_id and provincia_id and municipio_id):
        return HttpResponse('Debe proporcionar los parámetros departamento_id, provincia_id y municipio_id.', status=400)

    # Consulta SQL base con tablas intermedias
    sql_query = """
    SELECT r.id, p.codigo_sisin, rd.departamento_id, d.nombre AS departamento, rp.provincia_id, pr.nombre AS provincia, rm.municipio_id, m.nombre AS municipio,
           TRIM(p.nombre) AS nombre, p.programa_id, pro.sigla_prog_convenio AS programa,
           organizacion_id, org.sigla AS organizacion,
           p.ejecutor_id, e.nombre AS ejecutor, 
           tipo_id, tp.nombre AS tipo,
           p.estado_id, ep.nombre AS estado, TRIM(p.estado_detallado) AS estado_detallado,
           p.fecha_inicio, p.fecha_conclusion, r.total_inversion, r.avance_fisico, r.avance_financiamiento,
           p.gobierno  
    FROM proyectos p 
    INNER JOIN realizaciones r ON p.id = r.proyecto_id
    INNER JOIN realizacion_departamentos rd ON r.id = rd.realizacion_id
    INNER JOIN departamentos d ON rd.departamento_id = d.id
    INNER JOIN realizacion_provincias rp ON r.id = rp.realizacion_id
    INNER JOIN provincias pr ON rp.provincia_id = pr.id
    INNER JOIN realizacion_municipios rm ON r.id = rm.realizacion_id
    INNER JOIN municipios m ON rm.municipio_id = m.id
    INNER JOIN programas pro ON p.programa_id = pro.id
    INNER JOIN organizaciones_financieras org ON p.organizacion_id = org.id 
    INNER JOIN ejecutores e ON e.id = p.ejecutor_id  
    INNER JOIN tipos_proyecto tp ON tp.id = p.tipo_id  
    INNER JOIN estados2 ep ON ep.id = p.estado_id
    WHERE (%s IS NULL OR ep.id = %s)
    """

    params = [estado_id, estado_id]

    sql_query += " AND rd.departamento_id = %s"
    params.append(departamento_id)
    sql_query += " AND rp.provincia_id = %s"
    params.append(provincia_id)
    sql_query += " AND rm.municipio_id = %s"
    params.append(municipio_id)

    if realizacion_id:
        sql_query += " AND r.id = %s"
        params.append(realizacion_id)

    sql_query += " ORDER BY p.nombre;"

    with connection.cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()

        result_list = []
        for row in rows:
            result_dict = {
                "id": row[0],
                "codigo_sisin": row[1],
                "departamento_id": row[2],
                "departamento": row[3],
                "provincia_id": row[4],
                "provincia": row[5],
                "municipio_id": row[6],
                "municipio": row[7],
                "nombre": row[8],
                "programa_id": row[9],
                "sigla": row[10],
                "organizacion_id": row[11],
                "organizacion": row[12],
                "ejecutor_id": row[13],
                "ejecutor": row[14],
                "tipo_id": row[15],
                "tipo": row[16],
                "estado_id": row[17],
                "estado": row[18],
                "estado_detallado": row[19],
                "fecha_inicio": row[20],
                "fecha_conclusion": row[21],
                "total_inversion": row[22],
                "avance_fisico": row[23],
                "avance_financiamiento": row[24],
                "gobierno": row[25],
            }
            if row[17] == 3:
                proyectos.append(row[8])
                avances_fisicos.append(row[23])
                avances_financieros.append(row[24])
            result_list.append(result_dict)

    generar_grafico_barras_avances(proyectos, avances_fisicos, avances_financieros)
    context['first_row'] = result_list[0] if len(result_list) > 0 else None
    context['data'] = result_list
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sample.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Failed to generate PDF: {}'.format(pisa_status.err))

    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_general_pdf(request):
    departamento_id = request.GET.get('departamento_id')
    provincia_id = request.GET.get('provincia_id')
    municipio_id = request.GET.get('municipio_id')
    estado_id = request.GET.get('estado_id')

    resultado = ""
    total_inversion = 0
    total_programas = 0
    total_proyectos = 0
    
    lista_lugares = []
    numero_proyectos = []
    inversion = []

    # Consultar los registros de realizaciones filtrados según los parámetros recibidos
    realizaciones_queryset = Realizaciones.objects.all()

    if estado_id:
        realizaciones_queryset = realizaciones_queryset.filter(estado_id=estado_id)

    if municipio_id:
        realizaciones_queryset = realizaciones_queryset.filter(
            id__in=RealizacionMunicipios.objects.filter(municipio_id=municipio_id).values('realizacion_id')
        )
        resultado = realizaciones_queryset.values(
            nombre=F('municipio__nombre')
        ).annotate(
            cantidad_programas=Count('proyecto__programa_id', distinct=True),
            cantidad_proyectos=Count('proyecto_id', distinct=True),
            total_inversion=Sum('total_inversion')
        )

    elif provincia_id:
        realizaciones_queryset = realizaciones_queryset.filter(
            id__in=RealizacionProvincias.objects.filter(provincia_id=provincia_id).values('realizacion_id')
        )
        resultado = realizaciones_queryset.values(
            nombre=F('municipio__nombre')
        ).annotate(
            cantidad_programas=Count('proyecto__programa_id', distinct=True),
            cantidad_proyectos=Count('proyecto_id', distinct=True),
            total_inversion=Sum('total_inversion')
        )

    elif departamento_id:
        realizaciones_queryset = realizaciones_queryset.filter(
            id__in=RealizacionDepartamentos.objects.filter(departamento_id=departamento_id).values('realizacion_id')
        )
        resultado = realizaciones_queryset.values(
            nombre=F('provincia__nombre')
        ).annotate(
            cantidad_programas=Count('proyecto__programa_id', distinct=True),
            cantidad_proyectos=Count('proyecto_id', distinct=True),
            total_inversion=Sum('total_inversion')
        )

    else:
        resultado = realizaciones_queryset.values(
            nombre=F('departamento__nombre')
        ).annotate(
            cantidad_programas=Count('proyecto__programa_id', distinct=True),
            cantidad_proyectos=Count('proyecto_id', distinct=True),
            total_inversion=Sum('total_inversion')
        )
    for item in resultado:
        nombre_lugar = item.get('nombre')
        if nombre_lugar:
            lista_lugares.append(nombre_lugar)
        else:
            lista_lugares.append('otro')
    
        cantidad_proyectos = item.get('cantidad_proyectos', 0)
        numero_proyectos.append(cantidad_proyectos)
    
        total_inversion = item.get('total_inversion', 0)
        inversion_millones = round(float(total_inversion) / 1000000, 2) if total_inversion else 0
        inversion.append(inversion_millones)
    
        item['total_inversion'] = format_decimal(total_inversion, locale='es_BO', format='#,##0.00') if total_inversion else '0,00'

    
    grafico_barras_url = generar_grafico_barras_numero_proyectos(lista_lugares, numero_proyectos)
    grafico_barras2_url = generar_grafico_barras_inversion_proyectos(lista_lugares, inversion)
    total_inversion = realizaciones_queryset.aggregate(total_inversion=Sum('total_inversion'))['total_inversion']   
    total_proyectos = realizaciones_queryset.count()

    if total_proyectos:
        datos_grafico_tortas = dataGraficoTortas(departamento_id, provincia_id, municipio_id, estado_id)
    else:
        return HttpResponse({'No se encontraron proyectos en esa selección'}, status=status.HTTP_404_NOT_FOUND)

    sectores = datos_grafico_tortas['inversion_por_sector']
    datos_totales_inversion_torta = datos_grafico_tortas['total_inversion']
    datos_totales_proyectos_torta = datos_grafico_tortas['total_numero_proyectos']

    nombres_sectores = [sector['sector_nombre'] for sector in sectores]
    total_inversiones = [sector['total_inversion'] for sector in sectores]
    grafico_tortas = generar_grafico_tortas(nombres_sectores, total_inversiones)

    total_inversion_formateado = format_decimal(total_inversion, locale='es_ES', format='#,##0.00')

    template_path = 'reporte-general.html'
    base_url = f"{request.scheme}://{request.get_host()}"
    nombre_departamento = Departamento.objects.filter(id=departamento_id).first().nombre if departamento_id else None
    nombre_provincia = Provincia.objects.filter(id=provincia_id).first().nombre if provincia_id else None

    context = {
        'usuario':request.user,
        'base_url': base_url,
        'current_datetime': now(),
        'resultado': resultado,
        'total_proyectos': total_proyectos,
        'grafico_barras_url': grafico_barras_url,
        'grafico_barras2_url': grafico_barras2_url,
        'grafico_tortas_url': grafico_tortas,
        'datos_sectores': sectores,        
        'datos_totales_proyectos_torta': datos_totales_proyectos_torta,
        'datos_totales_inversion_torta': datos_totales_inversion_torta,
        'total_inversion': total_inversion_formateado,
        'nombre_departamento': nombre_departamento,
        'nombre_provincia': nombre_provincia
    }

    template = get_template(template_path)
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_general.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Failed to generate PDF: {}'.format(pisa_status.err))

    return response


def obtener_realizaciones(departamento_id=None, provincia_id=None, municipio_id=None, estado_id=None):
    realizaciones_queryset = Realizaciones.objects.all()

    if estado_id:
        realizaciones_queryset = realizaciones_queryset.filter(estado_id=estado_id)
    if municipio_id:
        realizaciones_queryset = realizaciones_queryset.filter(id__in=RealizacionMunicipios.objects.filter(municipio_id=municipio_id))
    elif provincia_id:
        realizaciones_queryset = realizaciones_queryset.filter(id__in=RealizacionProvincias.objects.filter(provincia_id=provincia_id))
    elif departamento_id:
        realizaciones_queryset = realizaciones_queryset.filter(id__in=RealizacionDepartamentos.objects.filter(departamento_id=departamento_id))

    return realizaciones_queryset

def dataGraficoTortas(departamento_id, provincia_id, municipio_id, estado_id):
    # Obtener las realizaciones según los filtros aplicados
    realizaciones_queryset = obtener_realizaciones(departamento_id, provincia_id, municipio_id, estado_id)
    
    # Agregar datos por sector
    agregacion_por_sector = realizaciones_queryset.values(sector_nombre=F('proyecto__sector__nombre')).annotate(
        total_inversion=Sum('total_inversion'),
        cantidad_proyectos=Count('proyecto_id')
    )
    
    # Calcular totales
    total_inversion_todos_los_sectores = realizaciones_queryset.aggregate(total_inversion=Sum('total_inversion'))['total_inversion'] 
    total_proyectos = realizaciones_queryset.count()

    # Formatear totales
    total_inversion_formateado = format_decimal(round(total_inversion_todos_los_sectores / 1000000, 2), locale='es_ES', format='#,##0.00')
    total_proyectos_formateado = format_decimal(total_proyectos, locale='es_ES', format='#,##0')

    # Formatear los datos agregados por sector
    for item in agregacion_por_sector:
        item['total_inversion'] = round(item['total_inversion'] / 1000000, 2)
    
    # Preparar la respuesta
    response_data = {
        'total_inversion': total_inversion_formateado,
        'total_numero_proyectos': total_proyectos_formateado,
        'inversion_por_sector': list(agregacion_por_sector)
    }
    
    return response_data


def generar_grafico_barras_numero_proyectos(lugares, n_proyectos):
    labels = lugares
    values = n_proyectos
    plt.figure(figsize=(10, 6))  # Aumenta el ancho para acomodar mejor los labels
    bars = plt.bar(labels, values, color='green')
    plt.xlabel('Lugares')
    plt.ylabel('Cantidad Proyectos')
    plt.title('NUMERO DE PROYECTOS POR LUGAR')
    # Rotar los labels del eje X a 45 grados
    plt.xticks(rotation=45, ha='right')
    # Añadir los valores del eje Y sobre cada barra
    for i, valor in enumerate(values):
        texto_etiqueta = f"{valor:.0f}"
        plt.text(i, valor + 0.1, texto_etiqueta, ha='center', va='bottom', fontsize=9)
        
    plt.tight_layout()
    plt.grid(axis='y', linestyle='--', linewidth=0.3)
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')    
    buffer.seek(0)    
    filename = 'grafico_barras.png'
    filepath = os.path.join(settings.MEDIA_ROOT, filename)
    with open(filepath, 'wb') as f:
        f.write(buffer.getbuffer())

    return os.path.join(settings.MEDIA_URL, filename)

def generar_grafico_barras_inversion_proyectos(lugares, inversion):
    labels = lugares
    values = inversion
    plt.figure(figsize=(10, 6))  # Aumenta el ancho para acomodar mejor los labels
    bars = plt.bar(labels, values, color='blue')
    plt.xlabel('Lugares')
    plt.ylabel('Inversion en Millones por LUGAR')
    plt.title('INVERSIONES en (M)')
    # Rotar los labels del eje X a 45 grados
    plt.xticks(rotation=45, ha='right')
    # Añadir los valores del eje Y sobre cada barra
    for i, valor in enumerate(values):
        texto_etiqueta = f"{valor:.2f}"
        plt.text(i, valor + 0.1, texto_etiqueta, ha='center', va='bottom', fontsize=8)
    plt.tight_layout()
    plt.grid(axis='y', linestyle='--', linewidth=0.3)
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')    
    buffer.seek(0)    
    filename = 'grafico_barras2.png'
    filepath = os.path.join(settings.MEDIA_ROOT, filename)
    with open(filepath, 'wb') as f:
        f.write(buffer.getbuffer())

    return os.path.join(settings.MEDIA_URL, filename)

def generar_grafico_tortas(nombres,valores):
    labels = nombres
    sizes = valores

    plt.figure(figsize=(6, 6))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
    plt.axis('equal')
    plt.title('% DE INVERSIÓN POR SECTORES')
    plt.tight_layout()
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    
    filename = 'grafico_torta.png'
    filepath = os.path.join(settings.MEDIA_ROOT, filename)
    with open(filepath, 'wb') as f:
        f.write(buffer.getbuffer())

    return os.path.join(settings.MEDIA_URL, filename)

import textwrap
def generar_grafico_barras_avances(proyectos, avances_fisicos, avances_financieros):
    # Envuelve los nombres de los proyectos para que se ajusten a múltiples líneas
    wrapped_labels = ['\n'.join(textwrap.wrap(label, 40)) for label in proyectos]

    avances_fisicos_values = avances_fisicos
    avances_financieros_values = avances_financieros

    y = range(len(wrapped_labels))
    height = 0.20  # Altura de las barras

    plt.figure(figsize=(12, 8))  # Ajusta el tamaño del gráfico
    bars2 = plt.barh([p + height for p in y], avances_financieros_values, height, label='% Avance Financiero', color='green')
    bars1 = plt.barh(y, avances_fisicos_values, height, label='% Avance Físico', color='blue')    

    plt.ylabel('PROYECTOS')
    plt.gca().yaxis.label.set_color('blue')
    plt.xlabel('PORCENTAJE DE AVANCE')
    plt.gca().xaxis.label.set_color('blue')
    plt.title('ESTADO DE AVANCE FÍSICO Y FINANCIERO DE PROYECTOS EN EJECUCIÓN')
    plt.yticks([p + height / 2 for p in y], wrapped_labels)
    plt.xticks(range(0, 101, 10))  # Marca de cada 10 hasta 100
    plt.xlim(0, 110)  # Establecer el límite del eje X a 100
    plt.legend()
    # Añadir los valores del eje X al final de cada barra
    for i, (av_fisico, av_financiero) in enumerate(zip(avances_fisicos_values, avances_financieros_values)):
        plt.text(av_fisico + 1, i, f"{av_fisico:.0f}%", va='center', fontsize=9)
        plt.text(av_financiero + 1, i + height, f"{av_financiero:.0f}%", va='center', fontsize=9)

    plt.tight_layout()
    plt.grid(axis='x', linestyle='--', linewidth=0.3)

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    filename = 'grafico_barras_avances.png'
    filepath = os.path.join(settings.MEDIA_ROOT, filename)
    with open(filepath, 'wb') as f:
        f.write(buffer.getbuffer())

    return os.path.join(settings.MEDIA_URL, filename)

from openpyxl import Workbook
def proyecto_por_municipio_excel(request):
    departamento_id = request.GET.get('departamento_id')
    provincia_id = request.GET.get('provincia_id')
    municipio_id = request.GET.get('municipio_id')
    estado_id = request.GET.get('estado_id')
    realizacion_id = request.GET.get('realizacion_id')

    if not (departamento_id and provincia_id and municipio_id):
        return HttpResponse('Debe proporcionar los parámetros departamento_id, provincia_id y municipio_id.', status=400)

    sql_query = """
    SELECT p.codigo_sisin, d.nombre AS departamento, pr.nombre AS provincia, m.nombre AS municipio,
            TRIM(p.nombre) AS nombre, pro.sigla_prog_convenio AS programa,
            org.sigla AS organizacion,
            e.nombre AS ejecutor, 
            tp.nombre AS tipo,
            ep.nombre AS estado, TRIM(p.estado_detallado) AS estado_detallado,
            p.fecha_inicio, p.fecha_conclusion, 
            regexp_replace(to_char(r.total_inversion, '999G999G999G999'), ',', '.', 'g') AS total_inversion, 
            r.avance_fisico, r.avance_financiamiento,
            p.gobierno  
    FROM proyectos p 
    INNER JOIN realizaciones r ON p.id = r.proyecto_id
    INNER JOIN realizacion_departamentos rd ON r.id = rd.realizacion_id
    INNER JOIN realizacion_provincias rp ON r.id = rp.realizacion_id
    INNER JOIN realizacion_municipios rm ON r.id = rm.realizacion_id
    INNER JOIN departamentos d ON rd.departamento_id = d.id
    INNER JOIN provincias pr ON rp.provincia_id = pr.id
    INNER JOIN municipios m ON rm.municipio_id = m.id
    INNER JOIN programas pro ON p.programa_id = pro.id
    INNER JOIN organizaciones_financieras org ON p.organizacion_id = org.id 
    INNER JOIN ejecutores e ON e.id = p.ejecutor_id  
    INNER JOIN tipos_proyecto tp ON tp.id = p.tipo_id  
    INNER JOIN estados2 ep ON ep.id = p.estado_id
    WHERE (%s IS NULL OR ep.id = %s)
    """

    params = [estado_id, estado_id]

    sql_query += " AND rd.departamento_id = %s"
    params.append(departamento_id)
    sql_query += " AND rp.provincia_id = %s"
    params.append(provincia_id)
    sql_query += " AND rm.municipio_id = %s"
    params.append(municipio_id)

    if realizacion_id:
        sql_query += " AND r.id = %s"
        params.append(realizacion_id)

    sql_query += " ORDER BY p.nombre;"

    with connection.cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Proyectos por Municipio"

    logo_path = os.path.join(settings.BASE_DIR, 'public', 'header-2.png')  # Ajusta el camino según la ubicación de tu logo
    logo_img = OpenpyxlImage(logo_path)
    logo_img.width = 500  # Ancho de la imagen en píxeles
    logo_img.height = 100  # Altura de la imagen en píxeles
    logo_img.anchor = 'A1'  # Posición en la hoja de cálculo
    worksheet.add_image(logo_img)

    # Ajustar la altura de la fila 1 para acomodar la imagen
    worksheet.row_dimensions[1].height = 100
    worksheet.merge_cells('A1:Q1')

    # Merge de celdas para el título
    worksheet.merge_cells('A2:Q2')
    cell = worksheet['A2']
    cell.value = "REPORTE DE PROYECTOS POR MUNICIPIO"
    cell.font = Font(size=20, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar el ancho de las columnas para el encabezado
    for col in range(1, 18):
        worksheet.column_dimensions[get_column_letter(col)].width = 20

    # Añadir encabezados de columnas
    headers = [
        "Código SISIN", "Departamento", "Provincia", 
        "Municipio", "Nombre", "Programa", 
        "Org. Financ.", "Ejecutor", "Tipo", "Estado", 
        "Estado Detallado", "Fecha Inicio", "Fecha Conclusión", "Total Inversión", 
        "Avance Físico", "Avance Financiamiento", "Gobierno"
    ]
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Color celeste bajito
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    worksheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        worksheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="proyectos_por_municipio.xlsx"'

    workbook.save(response)

    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_por_sector_y_estado_pdf(request):
    departamento_id = request.query_params.get('departamento_id') if request.query_params.get('departamento_id') !='null' else None
    provincia_id = request.query_params.get('provincia_id') if request.query_params.get('provincia_id')!='null' else None
    municipio_id = request.query_params.get('municipio_id') if request.query_params.get('municipio_id')!='null' else None
    estado_id = request.query_params.get('estado_id') if request.query_params.get('estado_id') !='null' else None
    sector_id = request.query_params.get('sector_id') if request.query_params.get('sector_id') !='null' else None
    
    filtro = {}

    if estado_id:
        filtro['estado_id'] = estado_id
    if municipio_id:
        filtro['realizacionmunicipios__municipio_id'] = municipio_id
    elif provincia_id:
        filtro['realizacionprovincias__provincia_id'] = provincia_id
    elif departamento_id:
        filtro['realizaciondepartamentos__departamento_id'] = departamento_id

    if sector_id:
        filtro['proyecto__sector_clasificador'] = sector_id

    realizaciones_queryset = Realizaciones.objects.filter(**filtro)


    if not filtro and not sector_id:
        realizaciones_queryset = Realizaciones.objects.all()


    # Obtener datos agregados por sector y estado
    resultados = (
        realizaciones_queryset.values('proyecto__sector_clasificador__nombre', 'estado__nombre')
        .annotate(
            nro_proyectos=Count('id'),
            total_inversion=Sum('total_inversion')
        )
        .order_by('proyecto__sector_clasificador__nombre', 'estado__nombre')
    )

    # Organizar los datos por sector
    data_por_sector = defaultdict(list)
    subtotales_por_sector = {}

    for resultado in resultados:
        sector = resultado['proyecto__sector_clasificador__nombre']
        data_por_sector[sector].append(resultado)
        if sector not in subtotales_por_sector:
            subtotales_por_sector[sector] = {'nro_proyectos': 0, 'total_inversion': 0}

        subtotales_por_sector[sector]['nro_proyectos'] += resultado['nro_proyectos']
        subtotales_por_sector[sector]['total_inversion'] += resultado['total_inversion'] or 0

    # Combinar los datos con subtotales
    data_con_subtotales = []
    for sector, datos in data_por_sector.items():
        data_con_subtotales.extend(datos)
        data_con_subtotales.append({
            'proyecto__sector_clasificador__nombre': sector,
            'estado__nombre': 'SUB TOTAL',
            'nro_proyectos': subtotales_por_sector[sector]['nro_proyectos'],
            'total_inversion': subtotales_por_sector[sector]['total_inversion'],
        })

    # Generar el PDF
    base_url = f"{request.scheme}://{request.get_host()}"
    template_path = 'reporte_por_sector_y_estado.html'
    context = {
        'usuario': request.user,
        'data': data_con_subtotales,
        'base_url': base_url,
        'current_datetime': now(),
        'departamento': Departamento.objects.filter(id=departamento_id).first().nombre if departamento_id  else None,
        'provincia':Provincia.objects.filter(id=provincia_id).first().nombre if provincia_id  else None,
        'municipio':Municipio.objects.filter(id=municipio_id).first().nombre if municipio_id  else None,
   
    }
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_general.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse(f'Failed to generate PDF: {pisa_status.err}', status=500)

    return response



